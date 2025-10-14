from typing import List, Dict, Optional, Any
import openpyxl 

# --- Classes de Entidade ---

class Peca:
    """Representa uma peça produzida com seus atributos e status de qualidade."""
    def __init__(self, id: str, peso: float, cor: str, comprimento: float):
        self.id = id
        self.peso = peso
        self.cor = cor.lower()
        self.comprimento = comprimento
        self.aprovada = False
        self.motivo_reprovacao: Optional[str] = None

    def __str__(self):
        status = "APROVADA" if self.aprovada else f"REPROVADA ({self.motivo_reprovacao})"
        return (f"ID: {self.id} | Peso: {self.peso}g | Cor: {self.cor.capitalize()} | "
                f"Comp.: {self.comprimento}cm | Status: {status}")

# --- Classe Principal de Controle ---

class ControleQualidade:
    """Gerencia o processo de inspeção, armazenamento e relatórios."""
    
    CAPACIDADE_CAIXA = 10

    def __init__(self):
        self.pecas_inspecionadas: Dict[str, Peca] = {}
        self.pecas_aprovadas: List[Peca] = []
        self.pecas_reprovadas: List[Peca] = []
        self.caixas_fechadas: List[List[Peca]] = []
        self.caixa_atual: List[Peca] = []

    def inspecionar_peca(self, peca: Peca) -> None:
        """
        Recebe uma peça, aplica os critérios de qualidade e a armazena.
        """
        if peca.id in self.pecas_inspecionadas:
            peca.motivo_reprovacao = "ID duplicado."
            self.pecas_reprovadas.append(peca)
            return

        motivos: List[str] = []

        # Critério 1: Peso
        if not (95 <= peca.peso <= 105):
            motivos.append(f"Peso ({peca.peso}g) fora da faixa (95g-105g).")

        # Critério 2: Cor
        if peca.cor not in ["azul", "verde"]:
            motivos.append(f"Cor ({peca.cor.capitalize()}) não é Azul ou Verde.")

        # Critério 3: Comprimento
        if not (10 <= peca.comprimento <= 20):
            motivos.append(f"Comp. ({peca.comprimento}cm) fora da faixa (10cm-20cm).")

        # Avaliação Final
        if not motivos:
            peca.aprovada = True
            self._adicionar_a_caixa(peca)
            self.pecas_aprovadas.append(peca)
        else:
            peca.aprovada = False
            peca.motivo_reprovacao = "; ".join(motivos)
            self.pecas_reprovadas.append(peca)

        self.pecas_inspecionadas[peca.id] = peca

    def _adicionar_a_caixa(self, peca: Peca) -> None:
        """Adiciona a peça aprovada à caixa atual e verifica se deve fechar."""
        self.caixa_atual.append(peca)
        if len(self.caixa_atual) == self.CAPACIDADE_CAIXA:
            self.caixas_fechadas.append(self.caixa_atual)
            self.caixa_atual = []
            print(f"\n--- CAIXA FECHADA --- (Capacidade: {self.CAPACIDADE_CAIXA} peças)")

    def remover_peca(self, peca_id: str) -> bool:
        """Remove uma peça pelo ID, se ela existir."""
        peca_id = peca_id.strip()
        
        if peca_id not in self.pecas_inspecionadas:
            return False

        peca = self.pecas_inspecionadas.pop(peca_id)
        
        # Remove das listas de status
        if peca.aprovada:
            self.pecas_aprovadas = [p for p in self.pecas_aprovadas if p.id != peca_id]
            # Tenta remover da caixa atual (se estiver lá)
            self.caixa_atual = [p for p in self.caixa_atual if p.id != peca_id]
            # Tenta remover das caixas fechadas 
            for i, caixa in enumerate(self.caixas_fechadas):
                 if any(p.id == peca_id for p in caixa):
                    print(f"ATENÇÃO: Peça {peca_id} removida de uma caixa FECHADA ({i+1}). A contagem de peças nesta caixa foi alterada.")
                    self.caixas_fechadas[i] = [p for p in caixa if p.id != peca_id]
        else:
            self.pecas_reprovadas = [p for p in self.pecas_reprovadas if p.id != peca_id]

        return True

    def gerar_relatorio(self) -> Dict[str, Any]:
        """Gera um dicionário com os dados consolidados do processo."""
        
        # Inclui a caixa atual no cálculo de caixas totais (se não estiver vazia)
        num_caixas_utilizadas = len(self.caixas_fechadas)
        if self.caixa_atual:
            num_caixas_utilizadas += 1

        # Consolida motivos de reprovação
        motivos_consolidados: Dict[str, int] = {}
        for peca in self.pecas_reprovadas:
            motivo = peca.motivo_reprovacao or "Motivo Desconhecido"
            motivos_consolidados[motivo] = motivos_consolidados.get(motivo, 0) + 1
        
        return {
            "total_aprovadas": len(self.pecas_aprovadas),
            "total_reprovadas": len(self.pecas_reprovadas),
            "motivos_reprovacao": motivos_consolidados,
            "quantidade_caixas_utilizadas": num_caixas_utilizadas,
            "pecas_na_caixa_atual": len(self.caixa_atual)
        }

    def gerar_relatorio_excel(self, nome_arquivo: str = "Relatorio_Producao.xlsx") -> None:
        """
        Gera um arquivo XLSX com o relatório consolidado e uma aba de detalhes.
        """
        dados = self.gerar_relatorio()
        
        # 1. Cria o Workbook e a primeira aba (Resumo)
        wb = openpyxl.Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo Geral"

        # Título
        ws_resumo.cell(row=1, column=1, value="RELATÓRIO CONSOLIDADO DE PRODUÇÃO E QUALIDADE")
        ws_resumo.merge_cells('A1:C1')

        # Dados Gerais
        ws_resumo.cell(row=3, column=1, value="TOTAL DE PEÇAS INSPECIONADAS:")
        ws_resumo.cell(row=3, column=2, value=len(self.pecas_inspecionadas))
        
        ws_resumo.cell(row=4, column=1, value="TOTAL DE PEÇAS APROVADAS:")
        ws_resumo.cell(row=4, column=2, value=dados['total_aprovadas'])
        
        ws_resumo.cell(row=5, column=1, value="TOTAL DE PEÇAS REPROVADAS:")
        ws_resumo.cell(row=5, column=2, value=dados['total_reprovadas'])

        # Dados de Caixas
        ws_resumo.cell(row=7, column=1, value="QUANTIDADE DE CAIXAS FECHADAS:")
        ws_resumo.cell(row=7, column=2, value=len(self.caixas_fechadas))
        
        ws_resumo.cell(row=8, column=1, value="TOTAL DE CAIXAS UTILIZADAS (FECHADAS + ATUAL):")
        ws_resumo.cell(row=8, column=2, value=dados['quantidade_caixas_utilizadas'])

        # Motivos de Reprovação
        ws_resumo.cell(row=10, column=1, value="MOTIVOS DE REPROVAÇÃO:")
        row = 11
        if dados['motivos_reprovacao']:
            for motivo, quantidade in dados['motivos_reprovacao'].items():
                ws_resumo.cell(row=row, column=1, value=motivo)
                ws_resumo.cell(row=row, column=2, value=quantidade)
                row += 1
        else:
            ws_resumo.cell(row=11, column=1, value="Nenhuma peça reprovada.")


        # 2. Cria a segunda aba (Detalhes das Peças)
        ws_detalhes = wb.create_sheet(title="Detalhes das Peças")
        
        # Cabeçalho da tabela de detalhes
        ws_detalhes.append(["ID", "Peso (g)", "Cor", "Comprimento (cm)", "Status", "Motivo Reprovação"])
        
        # Preenche os dados de todas as peças inspecionadas
        for peca_id in self.pecas_inspecionadas:
            peca = self.pecas_inspecionadas[peca_id]
            status = "APROVADA" if peca.aprovada else "REPROVADA"
            ws_detalhes.append([
                peca.id, 
                peca.peso, 
                peca.cor.capitalize(), 
                peca.comprimento, 
                status, 
                peca.motivo_reprovacao if peca.motivo_reprovacao else ""
            ])

        # 3. Salva o arquivo
        wb.save(nome_arquivo)